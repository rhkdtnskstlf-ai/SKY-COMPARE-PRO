import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

# 페이지 설정
st.set_page_config(layout="wide", page_title="Crew 분석 시스템 Pro")

# --- 커스텀 CSS ---
st.markdown("""
    <style>
    .stTable td, .stTable th {
        white-space: nowrap !important;
        font-size: 0.85rem !important;
        padding: 5px !important;
    }
    .group-card { background-color: #ffffff; border: 1px solid #e9ecef; border-left: 6px solid #fd7e14; padding: 20px; margin-bottom: 20px; border-radius: 12px; box-shadow: 2px 4px 12px rgba(0,0,0,0.08); }
    .move-group-card { background-color: #f0f4ff; border: 1px solid #dbe4ff; border-left: 6px solid #4c6ef5; padding: 20px; margin-bottom: 20px; border-radius: 12px; }
    .move-title { font-weight: 800; color: #1c7ed6; margin-bottom: 10px; font-size: 1.1em; }
    .flight-title { color: #1e293b; font-size: 1.3em; font-weight: 800; margin-bottom: 15px; border-bottom: 2px solid #f1f3f5; padding-bottom: 10px; }
    .item-container { display: flex; align-items: center; padding: 10px 14px; margin: 8px 0; border-radius: 8px; font-size: 0.95rem; line-height: 1.5; font-weight: 500; }
    .bg-swap { background-color: #f8f9fa; border-left: 4px solid #74c0fc; }
    .bg-in { background-color: #f3fcf3; border-left: 4px solid #40c057; }
    .bg-out { background-color: #fff5f5; border-left: 4px solid #ff8787; }
    .bg-info { background-color: #fff9db; border-left: 4px solid #fab005; }
    .badge { padding: 2px 8px; border-radius: 4px; font-size: 0.85em; font-weight: 800; margin-right: 12px; color: white; min-width: 45px; text-align: center; }
    .badge-swap { background-color: #1c7ed6; }
    .badge-in { background-color: #2f9e44; }
    .badge-out { background-color: #e03131; }
    .badge-info { background-color: #f08c00; }
    .flight-header { background-color: #f1f3f5; padding: 10px 15px; border-radius: 8px; font-weight: bold; margin-top: 15px; border-left: 4px solid #495057; }
    
    .user-guide {
        background-color: #f1f3f5;
        padding: 15px;
        border-radius: 10px;
        border-left: 5px solid #1c7ed6;
        margin-bottom: 20px;
    }
    .guide-title { font-weight: 800; color: #1c7ed6; font-size: 1rem; margin-bottom: 8px; }
    .guide-content { font-size: 0.85rem; color: #495057; line-height: 1.6; }
    .important { color: #e03131; font-weight: 800; text-decoration: underline; }
    </style>
    """, unsafe_allow_html=True)

# --- 유틸리티 함수 ---
def normalize_id(x):
    if pd.isna(x) or str(x).strip() == "": return ""
    text = str(x).strip()
    if text.endswith('.0'): text = text[:-2]
    try:
        return str(int(text))
    except ValueError:
        return text.replace(" ", "").upper()

def normalize_name(x):
    if pd.isna(x): return ""
    return str(x).replace(" ", "").replace("\t", "").strip().upper()

def format_time_display(val):
    if pd.isna(val) or val == "": return "-"
    if hasattr(val, 'strftime'): return val.strftime("%H:%M")
    try:
        str_val = str(val).strip()
        if len(str_val) >= 10: return pd.to_datetime(str_val).strftime("%H:%M")
        return str_val[:5]
    except: return str(val)

def is_layover_left(cell):
    try:
        if not cell.fill or cell.fill.fill_type is None: return False
        color_obj = cell.fill.start_color
        rgb_val = str(color_obj.rgb).upper() if color_obj.rgb else ""
        if rgb_val in ["", "00000000", "FFFFFFFF", "NONE"]: return False
        if "FFFF00" in rgb_val or "FFFFCC" in rgb_val or "FFFFE0" in rgb_val: return False
        return color_obj.index != 64
    except: return False

def get_status_icon(is_layover, dep_flt, forced_list):
    """로직 핵심: 리스트 포함 여부에 따라 아이콘 분기"""
    if str(dep_flt).strip().upper() in forced_list:
        return "🔴 3박"
    elif is_layover:
        return "✅"
    return ""

# --- 데이터 로더 ---
def load_crew_left(file, sheet_name):
    wb = openpyxl.load_workbook(file, data_only=False) 
    ws = wb[sheet_name]
    data = []
    for r in range(8, 500):
        c_id_val = ws[f"A{r}"].value
        if c_id_val is None: continue
        name_cell = ws[f"B{r}"] 
        raw_name = str(name_cell.value).strip() if name_cell.value else "Unknown"
        data.append({
            "CrewID": normalize_id(c_id_val),
            "CrewName": raw_name,
            "MatchName": normalize_name(raw_name),
            "Arr Flt": str(ws[f"G{r}"].value).strip().upper() if ws[f"G{r}"].value else "OPEN",
            "Arr Time": ws[f"H{r}"].value,
            "Dep Flt": str(ws[f"J{r}"].value).strip().upper() if ws[f"J{r}"].value else "OPEN", 
            "Dep Time": ws[f"K{r}"].value, 
            "is_layover": is_layover_left(name_cell),
            "Rank": ""
        })
    df = pd.DataFrame(data)
    df = df[df["CrewID"] != ""]
    return df

def load_crew_right(file):
    try:
        df = pd.read_excel(file, header=2, usecols="D:E,M:M,O:P,Q:R", engine='openpyxl')
        df.columns = ["CrewID", "CrewName", "Rank", "Arr Flt", "Arr Time", "Dep Flt", "Dep Time"]
        df["CrewID"] = df["CrewID"].apply(normalize_id)
        df["MatchName"] = df["CrewName"].apply(normalize_name)
        df["Date_Only"] = pd.to_datetime(df["Arr Time"], dayfirst=True, errors='coerce').dt.date
        df = df.dropna(subset=["CrewID"])
        return df
    except: return None

# --- 사이드바 ---
with st.sidebar:
    st.markdown("""
        <div class="user-guide">
            <div class="guide-title">💡 필독: 운영 및 아이콘 가이드</div>
            <div class="guide-content">
                1. <b>아이콘 구분:</b><br>
                - <b>✅ (초록 체크):</b> 일반 스테이/연박 인원<br>
                - <b>🔴 3박 (빨간 표시):</b> 지정된 특정 편수 인원<br><br>
                2. <b>작업 권장 사항:</b><br>
                - 기존 아날로그 방식과 로직이 다르므로, <span class="important">화면상의 변경 리포트(상단부터) 순서대로</span> 작업을 반영하시길 강력히 권장합니다.<br><br>
                3. <b>데이터 주의사항:</b><br>
                - 반드시 비교할 <b>NEW파일은 [기준일 + 1일]의 명단이 포함</b>되어 있어야 연박 분석이 정확합니다.<br><br>
                4. <b>피드백 및 오류 제보:</b><br>
                - 패턴 변경 시 며칠간 더블체크 부탁드리며, 특이사항은 <b>개발자</b>에게 전달 부탁드립니다. [바로 수정 예정]
            </div>
        </div>
    """, unsafe_allow_html=True)
    
    st.header("⚙️ 분석 설정")
    st.subheader("🏠 3박 편수 지정")
    forced_input = st.text_input("3박 처리할 편명 (쉼표 구분)", placeholder="예: KE123, KE456")
    forced_flts = [f.strip().upper() for f in forced_input.split(',')] if forced_input else []

    show_layover_only = st.checkbox("🏨 연박/3박 인원만 보기", value=False)
    
    if st.button("🔄 데이터 초기화"):
        st.cache_data.clear()
        st.rerun()

# --- 메인 분석 UI ---
st.title("✈️ Crew 명단 통합 분석 Pro")
up_l, up_r = st.columns(2)
df_l = df_r = None

with up_l:
    f_l = st.file_uploader("기존 명단 (Old)", type=["xlsx"])
    if f_l:
        wb_l = openpyxl.load_workbook(f_l, read_only=True)
        sh_l = st.selectbox("시트 선택", wb_l.sheetnames)
        df_l = load_crew_left(f_l, sh_l)

with up_r:
    f_r = st.file_uploader("신규 명단 (New)", type=["xlsx"])
    if f_r:
        df_r_raw = load_crew_right(f_r)
        if df_r_raw is not None:
            u_dates = sorted([d for d in df_r_raw["Date_Only"].unique() if pd.notna(d)])
            if u_dates:
                sel_d = st.selectbox("도착일 선택", u_dates)
                next_d = sel_d + timedelta(days=1)
                today_data = df_r_raw[df_r_raw["Date_Only"] == sel_d].copy()
                next_day_ids = set(df_r_raw[df_r_raw["Date_Only"] == next_d]["CrewID"])
                
                # 기본 연박 로직 (익일 여부)
                today_data["is_layover"] = today_data["CrewID"].apply(lambda x: x in next_day_ids)
                df_r = today_data

# --- 분석 엔진 ---
if df_l is not None and df_r is not None:
    # 필터링 로직: 일반 연박이거나 3박 지정 편수이거나
    if show_layover_only:
        df_l = df_l[(df_l['is_layover'] == True) | (df_l['Dep Flt'].isin(forced_flts))].copy()
        df_r = df_r[(df_r['is_layover'] == True) | (df_r['Dep Flt'].isin(forced_flts))].copy()

    st.divider()
    view_l, view_center, view_r = st.columns([1.2, 1.6, 1.2])

    def display_list(container, df, title):
        with container:
            st.subheader(title)
            if df.empty:
                st.write("데이터 없음")
                return
            sorted_df = df.sort_values(by=["Arr Time", "Arr Flt", "CrewName"])
            for flt in sorted_df["Arr Flt"].unique():
                gp = sorted_df[sorted_df["Arr Flt"] == flt]
                st.markdown(f"<div class='flight-header'>{flt} ({len(gp)}명)</div>", unsafe_allow_html=True)
                disp = gp.copy()
                # --- [수정 핵심] 조건부 아이콘 할당 ---
                disp["🏨"] = disp.apply(lambda row: get_status_icon(row['is_layover'], row['Dep Flt'], forced_flts), axis=1)
                disp["이름(ID)"] = disp["CrewName"] + "(" + disp["CrewID"] + ")"
                disp["도착"] = disp["Arr Time"].apply(format_time_display)
                disp["출발"] = disp["Dep Flt"].astype(str) + " (" + disp["Dep Time"].apply(format_time_display) + ")"
                st.table(disp[["🏨", "도착", "이름(ID)", "출발"]])

    display_list(view_l, df_l, "⬅️ 기존 명단")
    display_list(view_r, df_r, "➡️ 신규 명단")

    with view_center:
        st.markdown("<h2 style='text-align: center; margin-bottom: 25px;'>📋 통합 변경 리포트</h2>", unsafe_allow_html=True)
        
        all_merged = pd.merge(df_l, df_r, on="CrewID", how="outer", suffixes=('_old', '_new'))
        
        # 편수 이동 로직
        moved_crew = all_merged[all_merged['Arr Flt_old'].notna() & all_merged['Arr Flt_new'].notna() & (all_merged['Arr Flt_old'] != all_merged['Arr Flt_new'])].copy()
        processed_ids = set(moved_crew['CrewID'].tolist())
        
        if not moved_crew.empty:
            move_groups = moved_crew.groupby(['Arr Flt_old', 'Arr Flt_new'])
            for (old_f, new_f), group in move_groups:
                new_time = format_time_display(group['Arr Time_new'].iloc[0])
                names_html = " ".join([f"<span class='badge' style='background-color:#4c6ef5; display:inline-block; margin-bottom:5px;'>{n}({i}/{r})</span>" for n, i, r in zip(group['CrewName_new'], group['CrewID'], group['Rank_new'])])
                st.markdown(f"<div class='move-group-card'><div class='move-title'>🚚 편수 이동: {old_f} ➔ {new_f} <small>({new_time})</small></div>{names_html}</div>", unsafe_allow_html=True)

        # 주요 변경사항 리스트업
        all_flts_ordered = sorted(list(set(df_l["Arr Flt"].unique()) | set(df_r["Arr Flt"].unique())))
        for flt in all_flts_ordered:
            curr_old = df_l[(df_l["Arr Flt"] == flt) & (~df_l["CrewID"].isin(processed_ids))]
            curr_new = df_r[(df_r["Arr Flt"] == flt) & (~df_r["CrewID"].isin(processed_ids))]
            
            old_ids, new_ids = set(curr_old["CrewID"]), set(curr_new["CrewID"])
            out_ids, in_ids, stay_ids = old_ids - new_ids, new_ids - old_ids, old_ids & new_ids
            
            items_html = []
            rem_list, add_list = curr_old[curr_old["CrewID"].isin(out_ids)].to_dict('records'), curr_new[curr_new["CrewID"].isin(in_ids)].to_dict('records')
            
            # 교체/삭제/추가 로직
            match_cnt = min(len(rem_list), len(add_list))
            for _ in range(match_cnt):
                r, a = rem_list.pop(0), add_list.pop(0)
                items_html.append(f"<div class='item-container bg-swap'><span class='badge badge-swap'>교체</span> {r['CrewName']} ➔ <b>{a['CrewName']}({a['CrewID']}/{a['Rank']})</b></div>")
            for r in rem_list: items_html.append(f"<div class='item-container bg-out'><span class='badge badge-out'>CXL</span> {r['CrewName']}</div>")
            for a in add_list: items_html.append(f"<div class='item-container bg-in'><span class='badge badge-in'>IN</span> <b>{a['CrewName']}({a['CrewID']}/{a['Rank']})</b></div>")

            # 변경 사항 (시간, 3박 여부 등)
            for sid in stay_ids:
                o_r, n_r = curr_old[curr_old["CrewID"] == sid].iloc[0], curr_new[curr_new["CrewID"] == sid].iloc[0]
                sub = []
                
                # 아이콘 변화 감지
                status_old = get_status_icon(o_r['is_layover'], o_r['Dep Flt'], forced_flts)
                status_new = get_status_icon(n_r['is_layover'], n_r['Dep Flt'], forced_flts)
                if status_old != status_new:
                    sub.append(f"상태 변경: {status_old if status_old else '일반'} ➔ <b>{status_new if status_new else '일반'}</b>")
                
                if format_time_display(o_r['Arr Time']) != format_time_display(n_r['Arr Time']):
                    sub.append(f"시간: {format_time_display(o_r['Arr Time'])} ➔ {format_time_display(n_r['Arr Time'])}")
                
                if sub:
                    items_html.append(f"<div class='item-container bg-info'><span class='badge badge-info'>변경</span> <b>{n_r['CrewName']}({n_r['CrewID']}/{n_r['Rank']})</b>: {' / '.join(sub)}</div>")

            if items_html:
                st.markdown(f"<div class='group-card'><div class='flight-title'>✈️ {flt}</div>{''.join(items_html)}</div>", unsafe_allow_html=True)

    st.success(f"✅ 필터 적용 완료 (3박 지정: {len(forced_flts)}개 편수)")
else:
    st.info("💡 파일을 업로드하여 분석을 시작하세요.")
